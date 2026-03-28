"""
Lucas Correia
LIACS | Leiden University
Einsteinweg 55 | 2333 CC Leiden | The Netherlands
"""

import os

os.environ["CUDA_VISIBLE_DEVICES"] = "-1"
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"
import numpy as np
import tensorflow as tf
from tqdm.contrib import itertools
from dotenv import load_dotenv
from sklearn import metrics
from utilities import detection_class
import pandas as pd
import openpyxl

seeds = [1, 2, 3]
folds = [0, 1, 2]
splits = ["1day", "1week", "2weeks", "3weeks", "4weeks"]

# Declare constants
BUDGET = 10
MISLABEL_PROB = 0

# Load variables in .env file
load_dotenv()

# Load directory paths from .env file
data_path = os.path.join(os.environ["data_path"], "dqs")
model_path = os.path.join(os.environ["model_path"], "dqs")

results = []
for seed, fold_idx in itertools.product(seeds, folds):
    # Set fixed seed for random operations
    np.random.seed(seed)
    os.environ["PYTHONHASHSEED"] = str(seed)

    query_list = []
    query_score_list = []
    for split_idx, split in enumerate(splits):
        data_load_path = os.path.join(
            data_path, "2_preprocessed", "fold_" + str(fold_idx)
        )

        model_name = (
            "tevae_" + split + "_" + str(fold_idx) + "_1"
        )  # Fixed model seed due to focus on query strategy
        model_load_path = os.path.join(model_path, model_name)

        # Load tf validation data to extract window size
        tfdata_val = tf.data.Dataset.load(os.path.join(data_load_path, split, "val"))

        detector = detection_class.AnomalyDetector(
            model_path=model_load_path,
            window_size=tfdata_val.element_spec.shape[0],  # type: ignore
            sampling_rate=2,
            original_sampling_rate=10,
            calculate_delay=True,
            label_keyword="normal",
            mislabel_prob=MISLABEL_PROB,
            budget=BUDGET,
        )

        # Load input data
        train_list = detector.load_pickle(
            os.path.join(data_load_path, split, "train.pkl")
        )
        val_list = detector.load_pickle(os.path.join(data_load_path, split, "val.pkl"))
        test_list = detector.load_pickle(
            os.path.join(data_load_path, split, "test.pkl")
        )

        # Load inference data
        train_detection_score_list = detector.load_pickle(
            os.path.join(model_load_path, "train_detection_score.pkl")
        )
        val_detection_score_list = detector.load_pickle(
            os.path.join(model_load_path, "val_detection_score.pkl")
        )
        test_detection_score_list = detector.load_pickle(
            os.path.join(model_load_path, "test_detection_score.pkl")
        )

        # Combine training and validation data to get candidate list
        candidate_list = train_list + val_list
        candidate_score_list = train_detection_score_list + val_detection_score_list

        # Get initial threshold from candidate_score_list
        if len(query_score_list) == 0:
            threshold = detector.unsupervised_threshold(candidate_score_list)

        # Remove previously queried data from candidate list
        queried_filenames_list = [
            query_ts.dtype.metadata["file_name"] for query_ts in query_list
        ]
        candidate_list = [
            candidate_ts
            for candidate_ts in candidate_list
            if candidate_ts.dtype.metadata["file_name"] not in queried_filenames_list
        ]
        candidate_score_list = [
            candidate_score
            for candidate_score in candidate_score_list
            if candidate_score.dtype.metadata["file_name"] not in queried_filenames_list
        ]

        # Query strategy
        query_strategy_outputs = detector.uncertainty_query_strategy(
            query_list,
            query_score_list,
            candidate_list,
            candidate_score_list,
            threshold=threshold,
        )
        query_list, query_score_list, candidate_list, candidate_score_list = (
            query_strategy_outputs
        )

        # Extract groundtruth labels from query_list and test_list
        query_groundtruth_list = detector.extract_groundtruth(query_list)
        test_groundtruth_list = detector.extract_groundtruth(test_list)

        # Flip labels randomly with a given probability
        query_contaminated_list = detector.corrupt_labels(query_groundtruth_list)

        # Grid search for thresholds using labelled data
        f1_list = []
        reduced_query_score_list = np.concatenate(query_score_list).ravel()
        percentile_array = np.arange(0, 100.01, 0.01)
        for threshold_percentile in percentile_array:
            threshold_temp = np.percentile(
                reduced_query_score_list, threshold_percentile
            )
            groundtruth_labels_temp, predicted_labels_temp, _ = (
                detector.evaluate_online(
                    query_score_list,
                    threshold_temp,
                    groundtruth_list=query_contaminated_list,
                )
            )
            f1_list.append(
                metrics.f1_score(
                    groundtruth_labels_temp, predicted_labels_temp, zero_division=0
                )
            )
        f1_list = np.vstack(f1_list)
        threshold = np.percentile(
            reduced_query_score_list, percentile_array[np.argmax(f1_list)]
        )

        # Evaluate using threshold
        groundtruth_labels, predicted_labels, total_delays = detector.evaluate_online(
            test_detection_score_list,
            threshold,
            groundtruth_list=test_groundtruth_list,
        )

        results.append(
            {
                "Seed": seed,
                "Fold": fold_idx,
                "Split": split,
                "F1": metrics.f1_score(
                    groundtruth_labels, predicted_labels, zero_division=0
                ),
                "Precision": metrics.precision_score(
                    groundtruth_labels, predicted_labels, zero_division=0
                ),
                "Recall": metrics.recall_score(
                    groundtruth_labels, predicted_labels, zero_division=0
                ),
                "Delay": np.mean(total_delays),
                "Threshold": threshold,
            }
        )

results = pd.DataFrame(results)

if not os.path.isfile(os.path.join(model_path, "results.xlsx")):
    # Create and save a valid Excel file
    wb = openpyxl.Workbook()
    wb.save(os.path.join(model_path, "results.xlsx"))

# Use a try-finally block to ensure proper handling
try:
    with pd.ExcelWriter(
        os.path.join(model_path, "results.xlsx"), mode="a", if_sheet_exists="overlay"
    ) as writer:
        results.to_excel(
            writer, index=False, sheet_name=f"unc_{BUDGET}_{int(MISLABEL_PROB * 100)}"
        )
finally:
    # Cleanup: Remove default 'Sheet' if it exists
    try:
        workbook = openpyxl.load_workbook(os.path.join(model_path, "results.xlsx"))
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
        workbook.save(os.path.join(model_path, "results.xlsx"))
    except Exception as e:
        print(f"Error cleaning up sheets: {e}")
